import os
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment
from tempfile import NamedTemporaryFile
import argparse
import re


def get_column_letter(index):
    """
    Convert a zero-indexed column number to a column letter (A-Z, AA-AZ, BA-BZ, ..., ZZ), 0 -> A.
    """
    # Initialize an empty string to store the column letter
    column_letter = ""
    # Loop until the index is 0
    while index >= 0:
        # Calculate the remainder and the new index
        remainder = index % 26
        index = index // 26 - 1
        # Convert the remainder to a letter and prepend it to the column_letter string
        column_letter = chr(remainder + 65) + column_letter
    return column_letter


# 解析命令行参数
parser = argparse.ArgumentParser()
parser.add_argument(
    "--filter", type=str, help="Regex pattern to filter the description"
)
parser.add_argument("--per-row", type=int, default=3, help="Number of contents per row")
parser.add_argument(
    "--max-per-row", type=int, default=99, help="Maximum number of contents per row"
)
parser.add_argument(
    "--img-height", type=int, default=512, help="Height of embeded images by px"
)
parser.add_argument(
    "--input-dir", type=str, default="./output_selected/", help="Input dir"
)
parser.add_argument(
    "--output", type=str, default="spellbook.xlsx", help="Output file name"
)

args = parser.parse_args()
# 编译提供的正则表达式
if args.filter:
    pattern = re.compile(args.filter)
# 从参数来确定每行的内容数量
args = parser.parse_args()
contents_per_row = args.per_row
max_per_row = args.max_per_row

folder_path = args.input_dir
output_path = args.output
temp_folder = "tmp"
if not os.path.exists(temp_folder):
    os.mkdir(temp_folder)

img_px_height = args.img_height
img_cell_height = img_px_height * 3 / 4
img_cell_width = 0
ratio = 50 / 405
text_cell_width = 30

# 创建一个新的Excel工作簿
wb = Workbook()
ws = wb.active

# 临时文件路径列表
tmp_paths = []


def insert_img(img, row, col):
    # 等比缩放图片
    aspect_ratio = img.width / img.height
    new_height = img_px_height
    new_width = aspect_ratio * new_height
    img = img.resize((int(new_width), new_height), Image.Resampling.LANCZOS)
    # 使用NamedTemporaryFile创建一个临时文件
    with NamedTemporaryFile(delete=False, dir=temp_folder, suffix=".png") as tmp:
        img.save(tmp.name)
        temp_image_path = tmp.name
        tmp_paths.append(temp_image_path)
    # 插入图片，调整格式
    img_to_insert = OpenpyxlImage(temp_image_path)
    ws.add_image(img_to_insert, f"{get_column_letter(col)}{row}")
    ws.row_dimensions[row].height = img_cell_height
    return new_width


# 设置默认格式
for i in range(contents_per_row):
    ws.column_dimensions[get_column_letter(2 * i)].width = text_cell_width
alignment = Alignment(vertical="top", wrap_text=True)

col, row = 0, 1
description_history = {}
# 遍历文件夹中的所有文件
for filename in os.listdir(folder_path):
    if filename.lower().endswith((".png")):
        # 获取图片完整路径
        image_path = os.path.join(folder_path, filename)

        # 打开图片并读取元信息
        with Image.open(image_path) as img:
            description = img.info.get("Description", "无描述信息")
            if args.filter:
                description = ",".join(
                    [d for d in description.split(",") if pattern.search(d)]
                ).strip()
            if description == "":
                continue

            # 当每行内容大于1时，按顺序插入metadata和图片
            if contents_per_row != 1:
                # 在当前列写入'Description'元信息
                cell = f"{get_column_letter(col)}{row}"
                ws[cell] = description
                ws[cell].alignment = alignment
                # 插入图片到下一列
                img_cell_width = max(img_cell_width, insert_img(img, row, col + 1))
                # 移动到下个单元格
                col += 2
                if col == (contents_per_row) * 2:
                    row += 1
                    col = 0

            # 当每行内容为1时，将相同metadata对应的图片归为同一列
            else:
                row, col = description_history.get(
                    description, (len(description_history) + 1, 0)
                )
                if col > max_per_row:
                    continue
                # 如果之前没有，则写入metadata
                if col == 0:
                    cell = f"{get_column_letter(col)}{row}"
                    ws[cell] = description
                    ws[cell].alignment = alignment
                    col += 1
                # 插入图片
                img_cell_width = max(img_cell_width, insert_img(img, row, col))
                # 记录插入的metadata历史
                description_history[description] = (row, col + 1)

img_cell_width = int(img_cell_width * ratio)
if contents_per_row != 1:
    for i in range(contents_per_row):
        img_col = get_column_letter(i * 2 + 1)
        ws.column_dimensions[img_col].width = img_cell_width
else:
    max_col = 0
    for key, (row_num, col_num) in description_history.items():
        max_col = max(max_col, col_num)
    for i in range(1, max_col):
        ws.column_dimensions[get_column_letter(i)].width = img_cell_width

# 保存Excel文件
wb.save(output_path)

# 删除临时文件
for path in tmp_paths:
    if os.path.exists(path):
        os.remove(path)
